from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
SPEC = importlib.util.spec_from_file_location("eda_test_plan_validator", ROOT / "scripts" / "test_plan_validator.py")
if SPEC is None or SPEC.loader is None:
    raise RuntimeError("cannot load test-plan validator")
VALIDATOR = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = VALIDATOR
SPEC.loader.exec_module(VALIDATOR)


class TestPlanValidatorTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def workbook(self, name: str = "plan.xlsx", values: list[object] | None = None) -> Path:
        path = self.root / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "cmd"
        sheet.append(list(VALIDATOR.HEADERS))
        sheet.append(values or [
            1,
            "add_net",
            "cases/add_net/pos001_minimum",
            "合法最小输入场景，覆盖 add_net 的 name 选项",
            "1. 准备空设计。\n2. 调用 add_net -name n1。\n3. 查询并收集网络对象。",
            "1. 命令无报错。\n2. 查询到完整的 n1 网络对象。",
            "测试员",
            "PASS",
            "覆盖 name 必选测试点",
            "EDA-1",
        ])
        workbook.save(path)
        workbook.close()
        return path

    def rules(self, path: Path, requirements: Path | None = None) -> list[str]:
        return [item.rule_id for item in VALIDATOR.validate_test_plan(path, requirements)]

    def assert_rule_passes(self, rule: str) -> None:
        self.assertNotIn(rule, self.rules(self.workbook()))

    def test_each_rule_has_a_passing_sample(self) -> None:
        path = self.workbook()
        requirements = self.root / "requirements.json"
        requirements.write_text(json.dumps({"add_net": ["name"]}), encoding="utf-8")
        rules = self.rules(path, requirements)
        for index in range(1, 15):
            with self.subTest(rule=index):
                self.assertNotIn(f"PLAN-{index:03d}", rules)

    def test_plan_001_rejects_wrong_extension_and_corrupt_xlsx(self) -> None:
        wrong = self.root / "plan.xls"
        wrong.write_text("not excel", encoding="utf-8")
        self.assertIn("PLAN-001", self.rules(wrong))
        corrupt = self.root / "plan.xlsx"
        corrupt.write_text("not excel", encoding="utf-8")
        self.assertIn("PLAN-001", self.rules(corrupt))

    def test_plan_002_requires_cmd_sheet(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        workbook["cmd"].title = "cases"
        workbook.save(path)
        workbook.close()
        self.assertIn("PLAN-002", self.rules(path))

    def test_plan_003_requires_headers_and_empty_extra_columns(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        sheet = workbook["cmd"]
        sheet["A1"] = "id"
        sheet["K2"] = "undefined"
        workbook.save(path)
        workbook.close()
        self.assertIn("PLAN-003", self.rules(path))

    def test_plan_004_and_014_reject_partial_rows(self) -> None:
        values = [1, "add_net", "cases/add_net/pos001_minimum", "", "", "", "", "", "", ""]
        rules = self.rules(self.workbook(values=values))
        self.assertIn("PLAN-004", rules)
        self.assertIn("PLAN-014", rules)

    def test_plan_005_requires_positive_integer(self) -> None:
        values = [0, "add_net", "cases/add_net/pos001_minimum", "合法最小输入场景，覆盖 name 选项", "1. 准备设计。\n2. 调用 add_net。\n3. 检查对象。", "查询对象完整", "", "", "", ""]
        self.assertIn("PLAN-005", self.rules(self.workbook(values=values)))

    def test_plan_006_requires_one_command_name(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        workbook["cmd"]["B2"] = "add_net delete_net"
        workbook.save(path)
        workbook.close()
        self.assertIn("PLAN-006", self.rules(path))

    def test_plan_007_checks_case_name_duplicates_and_indices(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        sheet = workbook["cmd"]
        sheet["C2"] = "cases/add_net/pos002_bad"
        sheet.append(list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0])
        workbook.save(path)
        workbook.close()
        self.assertIn("PLAN-007", self.rules(path))

    def test_plan_008_requires_meaningful_description(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        workbook["cmd"]["D2"] = "功能测试"
        workbook.save(path)
        workbook.close()
        self.assertIn("PLAN-008", self.rules(path))

    def test_plan_009_requires_numbered_actionable_steps(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        workbook["cmd"]["E2"] = "运行一下"
        workbook.save(path)
        workbook.close()
        self.assertIn("PLAN-009", self.rules(path))

    def test_plan_010_requires_observable_expected_result(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        workbook["cmd"]["F2"] = "PASS"
        workbook.save(path)
        workbook.close()
        self.assertIn("PLAN-010", self.rules(path))

    def test_plan_011_reports_warning_without_failing_layer(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        workbook["cmd"]["G2"] = "TBD"
        workbook["cmd"]["H2"] = "UNKNOWN"
        workbook.save(path)
        workbook.close()
        diagnostics = VALIDATOR.validate_test_plan(path)
        warnings = [item for item in diagnostics if item.rule_id == "PLAN-011"]
        self.assertTrue(warnings)
        self.assertTrue(all(item.severity == "warning" for item in warnings))
        self.assertEqual("PASS", VALIDATOR.plan_report(path)["status"])

    def test_plan_012_rejects_polarity_mismatch(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        workbook["cmd"]["D2"] = "非法缺失 name 输入场景"
        workbook["cmd"]["F2"] = "预期错误并拒绝创建网络对象"
        workbook.save(path)
        workbook.close()
        self.assertIn("PLAN-012", self.rules(path))

    def test_plan_013_checks_requirements_manifest(self) -> None:
        requirements = self.root / "requirements.json"
        requirements.write_text(json.dumps({"add_net": ["name", "physical"]}), encoding="utf-8")
        self.assertIn("PLAN-013", self.rules(self.workbook(), requirements))

    def test_plan_014_rejects_hidden_case_row(self) -> None:
        path = self.workbook()
        from openpyxl import load_workbook
        workbook = load_workbook(path)
        workbook["cmd"].row_dimensions[2].hidden = True
        workbook.save(path)
        workbook.close()
        self.assertIn("PLAN-014", self.rules(path))


if __name__ == "__main__":
    unittest.main()
