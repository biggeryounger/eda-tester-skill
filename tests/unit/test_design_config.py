from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.design_config import (
    ConfigError,
    create_management_workbook,
    load_profile,
    render_design_tcl,
    sync_design_tcl,
)


class DesignConfigTests(unittest.TestCase):
    def profile(self) -> dict[str, object]:
        return {
            "schema_version": "1.0",
            "profiles": {
                "smoke": {
                    "design_dir": "$env(PV_ROOT)/designs/smoke",
                    "tech_dir": "$env(PV_ROOT)/tech/smoke",
                    "top_cell": "smoke_top",
                    "lef_files": ["$tech_dir/core.lef", "$tech_dir/io.lef"],
                    "netlist_file": "$design_dir/top.v.gz",
                    "def_file": "$design_dir/top.def.gz",
                    "power_net": "VDD",
                    "ground_net": "VSS",
                }
            },
        }

    def test_one_profile_renders_all_design_paths(self) -> None:
        profile = load_profile(self.profile(), "smoke")
        text = render_design_tcl(profile, "smoke")
        self.assertIn("# Generated from central design profile: smoke", text)
        self.assertIn('set design_dir "$env(PV_ROOT)/designs/smoke"', text)
        self.assertIn('"$tech_dir/core.lef"', text)
        self.assertIn('set netlist_file "$design_dir/top.v.gz"', text)
        self.assertIn('set def_file "$design_dir/top.def.gz"', text)

    def test_rejects_machine_specific_absolute_path(self) -> None:
        data = self.profile()
        data["profiles"]["smoke"]["design_dir"] = "/Users/name/design"
        with self.assertRaisesRegex(ConfigError, "portable"):
            load_profile(data, "smoke")

    def test_rejects_unknown_and_missing_fields(self) -> None:
        data = self.profile()
        del data["profiles"]["smoke"]["def_file"]
        data["profiles"]["smoke"]["netlist"] = "wrong"
        with self.assertRaisesRegex(ConfigError, "missing.*def_file.*unknown.*netlist"):
            load_profile(data, "smoke")

    def test_sync_updates_multiple_targets_identically_and_rejects_duplicates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            targets = [root / "pos001_a/tcl/design.tcl", root / "pos002_b/tcl/design.tcl"]
            sync_design_tcl(self.profile(), "smoke", targets)
            self.assertEqual(targets[0].read_text(), targets[1].read_text())
            with self.assertRaisesRegex(ConfigError, "duplicate output"):
                sync_design_tcl(self.profile(), "smoke", [targets[0], targets[0]])

    def test_repository_config_is_valid_json_and_has_example_profile(self) -> None:
        path = Path(__file__).resolve().parents[2] / "assets/design-profiles.json"
        data = json.loads(path.read_text(encoding="utf-8"))
        profile = load_profile(data, "repository_example")
        self.assertEqual("riscv_core", profile["top_cell"])

    def test_management_workbook_lists_every_central_input(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "输入件管理表.xlsx"
            create_management_workbook(self.profile(), output)
            wb = load_workbook(output, data_only=False)
            ws = wb["输入件管理"]
            self.assertEqual(
                ["编号", "Profile", "输入类别", "变量/字段", "路径或值", "必选", "状态", "冲突检查", "来源", "备注"],
                [cell.value for cell in ws[1]],
            )
            self.assertEqual(9, ws.max_row)
            self.assertEqual("$env(PV_ROOT)/designs/smoke", ws[2][4].value)
            self.assertEqual("有效", ws[2][6].value)
            self.assertEqual("通过", ws[2][7].value)
            self.assertEqual("assets/design-profiles.json", ws[2][8].value)
            self.assertEqual("输入说明", wb.sheetnames[1])
            wb.close()


if __name__ == "__main__":
    unittest.main()
