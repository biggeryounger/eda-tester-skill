from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
GENERATOR = ROOT / "scripts" / "generate_test_plan.py"


class TestPlanGeneratorTests(unittest.TestCase):
    def test_template_excludes_the_two_unwanted_note_images_only(self) -> None:
        template = ROOT / "assets" / "templates" / "测试用例设计表.xlsx"
        with zipfile.ZipFile(template) as archive:
            drawing = ET.fromstring(archive.read("xl/drawings/drawing1.xml"))
        ns = {"xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
        anchors = []
        for anchor in drawing.findall("xdr:twoCellAnchor", ns):
            origin = anchor.find("xdr:from", ns)
            anchors.append(
                (
                    int(origin.find("xdr:row", ns).text),
                    int(origin.find("xdr:col", ns).text),
                )
            )
        self.assertEqual(7, len(anchors))
        self.assertNotIn((1, 8), anchors)
        self.assertNotIn((2, 10), anchors)
        self.assertIn((5, 10), anchors)

    def test_skill_delivery_contract_excludes_case_tree_and_requires_generator(self) -> None:
        skill = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("generate_test_plan.py", skill)
        self.assertIn("Do not hand off an empty workbook", skill)
        self.assertNotIn("Create one `case.toml`", skill)
        self.assertNotIn("Generate a real case tree", skill)
        self.assertNotIn("用例目录与运行说明", skill)

    def test_generator_writes_nonempty_valid_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            source = root / "scenarios.json"
            output = root / "plan.xlsx"
            source.write_text(json.dumps({"scenarios": [
                {
                    "number": 1,
                    "被测命令": "add_net",
                    "用例路径": "pos001_minimum",
                    "用例描述": "[feature:F-003] [strategy:smoke] 合法最小调用场景，覆盖 add_net 的 name 选项",
                    "用例步骤": "1. 准备空设计。\n2. 调用 add_net -name n1。\n3. 查询并收集网络对象。",
                    "用例预期": "1. 命令无报错。\n2. 查询到完整的 n1 网络对象。",
                    "责任人": "",
                    "状态": "NOT_RUN",
                    "备注": "策略：Smoke case",
                    "Ticket": "",
                }
            ]}, ensure_ascii=False), encoding="utf-8")
            completed = subprocess.run([sys.executable, str(GENERATOR), str(source), str(output)], cwd=ROOT, capture_output=True, text=True)
            self.assertEqual(0, completed.returncode, completed.stderr)
            workbook = load_workbook(output, data_only=False)
            sheet = workbook["cmd"]
            self.assertEqual(2, sheet.max_row)
            self.assertEqual("add_net", sheet["B2"].value)
            self.assertEqual("pos001_minimum", sheet["C2"].value)
            self.assertEqual("[strategy:smoke] 合法最小调用场景，覆盖 add_net 的 name 选项", sheet["D2"].value)
            self.assertNotIn("[feature", sheet["D2"].value.lower())
            self.assertFalse(sheet.row_dimensions[2].hidden)
            self.assertFalse(sheet.row_dimensions[2].collapsed)
            self.assertEqual("A1:J2", sheet.auto_filter.ref)
            expected_widths = {
                "A": 8,
                "B": 18,
                "C": 26,
                "D": 32,
                "E": 36,
                "F": 32,
                "G": 12,
                "H": 12,
                "I": 26,
                "J": 18,
            }
            for column, width in expected_widths.items():
                self.assertEqual(width, sheet.column_dimensions[column].width)
            for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=10):
                for cell in row:
                    self.assertTrue(cell.alignment.wrap_text)
            self.assertGreaterEqual(sheet.row_dimensions[2].height, 42)
            self.assertLessEqual(sheet.row_dimensions[2].height, 240)
            workbook.close()
            validation = subprocess.run([sys.executable, "scripts/validate.py", "plan", str(output)], cwd=ROOT, capture_output=True, text=True)
            self.assertEqual(0, validation.returncode, validation.stderr)

    def test_generator_rejects_empty_scenarios(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            source = root / "empty.json"
            output = root / "plan.xlsx"
            source.write_text('{"scenarios": []}', encoding="utf-8")
            completed = subprocess.run([sys.executable, str(GENERATOR), str(source), str(output)], cwd=ROOT, capture_output=True, text=True)
            self.assertNotEqual(0, completed.returncode)
            self.assertIn("at least one scenario", completed.stderr)
            self.assertFalse(output.exists())


if __name__ == "__main__":
    unittest.main()
