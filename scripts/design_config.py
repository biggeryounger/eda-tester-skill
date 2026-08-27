#!/usr/bin/env python3
"""Render case-local design.tcl files from one validated design profile."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


FIELDS = (
    "design_dir",
    "tech_dir",
    "top_cell",
    "lef_files",
    "netlist_file",
    "def_file",
    "power_net",
    "ground_net",
)
PROFILE_NAME = re.compile(r"^[a-z][a-z0-9_]*$")
MACHINE_PATH = re.compile(r"^(?:/Users/|/home/|[A-Za-z]:[\\/])")


class ConfigError(ValueError):
    """A deterministic design-profile validation error."""


def _strings(value: object) -> Iterable[str]:
    if isinstance(value, str):
        yield value
    elif isinstance(value, list):
        for item in value:
            if isinstance(item, str):
                yield item


def load_profile(data: Mapping[str, Any], profile_name: str) -> dict[str, Any]:
    if data.get("schema_version") != "1.0":
        raise ConfigError("schema_version must be '1.0'")
    if not PROFILE_NAME.fullmatch(profile_name):
        raise ConfigError(f"invalid profile name: {profile_name}")
    profiles = data.get("profiles")
    if not isinstance(profiles, dict) or profile_name not in profiles:
        raise ConfigError(f"unknown profile: {profile_name}")
    raw = profiles[profile_name]
    if not isinstance(raw, dict):
        raise ConfigError(f"profile {profile_name} must be an object")
    missing = sorted(set(FIELDS) - raw.keys())
    unknown = sorted(set(raw) - set(FIELDS))
    problems = []
    if missing:
        problems.append("missing fields: " + ", ".join(missing))
    if unknown:
        problems.append("unknown fields: " + ", ".join(unknown))
    if problems:
        raise ConfigError("; ".join(problems))
    for field in FIELDS:
        value = raw[field]
        if field == "lef_files":
            if not isinstance(value, list) or not value or not all(isinstance(item, str) and item.strip() for item in value):
                raise ConfigError("lef_files must be a non-empty string array")
        elif not isinstance(value, str) or not value.strip():
            raise ConfigError(f"{field} must be a non-empty string")
        for path_value in _strings(value):
            if MACHINE_PATH.search(path_value.strip()):
                raise ConfigError(f"{field} must use a portable PV_ROOT-anchored or relative path")
    return {field: raw[field] for field in FIELDS}


def _quote(value: str) -> str:
    return '"' + value.replace('"', '\\"') + '"'


def render_design_tcl(profile: Mapping[str, Any], profile_name: str) -> str:
    lef_separator = " \\\n    "
    lef_lines = lef_separator.join(_quote(item) for item in profile["lef_files"])
    return (
        f"# Generated from central design profile: {profile_name}\n"
        "# Edit assets/design-profiles.json and resync; do not edit paths here.\n"
        f"set design_dir {_quote(profile['design_dir'])}\n"
        f"set tech_dir {_quote(profile['tech_dir'])}\n\n"
        f"set init_top_cell {_quote(profile['top_cell'])}\n"
        f"set lef_files [list \\\n    {lef_lines}]\n"
        f"set netlist_file {_quote(profile['netlist_file'])}\n"
        f"set def_file {_quote(profile['def_file'])}\n"
        f"set power_net {_quote(profile['power_net'])}\n"
        f"set ground_net {_quote(profile['ground_net'])}\n"
    )


def sync_design_tcl(data: Mapping[str, Any], profile_name: str, targets: Iterable[Path]) -> list[Path]:
    target_list = [Path(target).resolve() for target in targets]
    if not target_list:
        raise ConfigError("at least one output target is required")
    if len(set(target_list)) != len(target_list):
        raise ConfigError("duplicate output target")
    text = render_design_tcl(load_profile(data, profile_name), profile_name)
    for target in target_list:
        if target.name != "design.tcl":
            raise ConfigError(f"output target must be named design.tcl: {target}")
    for target in target_list:
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(text, encoding="utf-8")
    return target_list


def create_management_workbook(data: Mapping[str, Any], output: Path) -> Path:
    profiles = data.get("profiles")
    if not isinstance(profiles, dict) or not profiles:
        raise ConfigError("profiles must be a non-empty object")
    wb = Workbook()
    ws = wb.active
    ws.title = "输入件管理"
    headers = ["编号", "Profile", "输入类别", "变量/字段", "路径或值", "必选", "状态", "冲突检查", "来源", "备注"]
    ws.append(headers)
    categories = {
        "design_dir": "目录", "tech_dir": "目录", "top_cell": "设计对象",
        "lef_files": "物理输入", "netlist_file": "逻辑输入", "def_file": "物理输入",
        "power_net": "网络", "ground_net": "网络",
    }
    required = {"design_dir", "tech_dir", "top_cell", "lef_files", "netlist_file", "def_file"}
    index = 1
    for profile_name in sorted(profiles):
        profile = load_profile(data, profile_name)
        for field in FIELDS:
            value = profile[field]
            display = "\n".join(value) if isinstance(value, list) else value
            ws.append([
                f"IN-{index:03d}", profile_name, categories[field], field, display,
                "是" if field in required else "否", "有效", "通过",
                "assets/design-profiles.json", "由集中配置自动生成，请勿在 design.tcl 中手工修改路径",
            ])
            index += 1
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [12, 22, 14, 20, 72, 10, 12, 14, 30, 52]
    for column, width in zip("ABCDEFGHIJ", widths):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    table = Table(displayName="DesignInputs", ref=f"A1:J{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    notes = wb.create_sheet("输入说明")
    notes.append(["项目", "说明"])
    notes.append(["唯一配置源", "assets/design-profiles.json"])
    notes.append(["更新流程", "修改 JSON profile → 生成管理表 → 批量同步各用例 design.tcl → 执行功能验证"])
    notes.append(["路径规则", "使用 $env(PV_ROOT) 锚定路径或相对路径；禁止 /Users、/home 和 Windows 用户盘绝对路径"])
    notes.append(["冲突规则", "同一次同步禁止重复输出目标；未知字段、缺失字段和不便携路径均拒绝生成"])
    for cell in notes[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 100
    notes.freeze_panes = "A2"
    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    wb.close()
    return output


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=Path("assets/design-profiles.json"))
    parser.add_argument("--profile", required=True)
    parser.add_argument("--management-workbook", type=Path)
    parser.add_argument("targets", nargs="*", type=Path)
    args = parser.parse_args()
    try:
        data = json.loads(args.config.read_text(encoding="utf-8"))
        written = sync_design_tcl(data, args.profile, args.targets) if args.targets else []
        if args.management_workbook:
            create_management_workbook(data, args.management_workbook)
        if not written and not args.management_workbook:
            raise ConfigError("provide design.tcl targets or --management-workbook")
    except (OSError, json.JSONDecodeError, ConfigError) as exc:
        parser.error(str(exc))
    for path in written:
        print(path)
    if args.management_workbook:
        print(args.management_workbook.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
