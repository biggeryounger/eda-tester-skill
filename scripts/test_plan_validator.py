#!/usr/bin/env python3
"""Deterministic Layer 1 validator for EDA test-plan workbooks."""

from __future__ import annotations

import json
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


HEADERS = ("number", "被测命令", "用例路径", "用例描述", "用例步骤", "用例预期", "责任人", "状态", "备注", "Ticket")
CASE_NAME = re.compile(r"^(pos|neg)([0-9]{3})_([a-z0-9][a-z0-9_]*)$")
COMMAND_NAME = re.compile(r"^[A-Za-z_][A-Za-z0-9_:.-]*$")
STEP_NUMBER = re.compile(r"(?m)^\s*1[.、)]")
ORDERED_ITEM_MARKER = re.compile(r"(?<!\S)(\d+)[.、]\s+")
OBSERVABLE = re.compile(r"报错|错误|warning|warn|拒绝|报告|文件|对象|数据|golden|iTools|比较|查询|生成|不存在|不变|完整", re.I)
NEGATIVE = re.compile(r"非法|缺失|冲突|不支持|错误|报错|warning|warn|拒绝|失败|不存在|超出|无效", re.I)
POSITIVE = re.compile(r"合法|支持|正常|典型|基本|有效|成功|无报错|完整|一致", re.I)
VALID_STATUSES = {"", "PASS", "FAIL", "BLOCKED", "TODO", "NOT_RUN", "SKIPPED"}


@dataclass(frozen=True)
class PlanDiagnostic:
    rule_id: str
    severity: str
    file: Path
    sheet: str | None
    row: int | None
    message: str

    def to_dict(self) -> dict[str, object]:
        return {
            "layer": "L1",
            "rule_id": self.rule_id,
            "severity": self.severity,
            "file": str(self.file),
            "sheet": self.sheet,
            "line": self.row,
            "message": self.message,
        }


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _diag(path: Path, rule: str, message: str, row: int | None = None, severity: str = "error") -> PlanDiagnostic:
    return PlanDiagnostic(rule, severity, path, "cmd" if row is not None else None, row, message)


def _has_inline_numbered_items(value: str) -> bool:
    matches = list(ORDERED_ITEM_MARKER.finditer(value))
    if len(matches) < 2:
        return False
    numbers = [int(match.group(1)) for match in matches]
    if numbers[0] != 1 or numbers != list(range(1, len(numbers) + 1)):
        return False
    return any("\n" not in value[previous.start():current.start()] for previous, current in zip(matches, matches[1:]))


def load_requirements(path: Path | None) -> tuple[dict[str, list[str]], list[PlanDiagnostic]]:
    if path is None:
        return {}, []
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return {}, [PlanDiagnostic("PLAN-013", "error", path, None, None, f"cannot read requirements: {exc}")]
    if not isinstance(raw, dict) or not all(
        isinstance(command, str) and isinstance(points, list) and all(isinstance(point, str) and point.strip() for point in points)
        for command, points in raw.items()
    ):
        return {}, [PlanDiagnostic("PLAN-013", "error", path, None, None, "requirements must map commands to non-empty string arrays")]
    return {command: [point.strip() for point in points] for command, points in raw.items()}, []


def validate_test_plan(path: Path, requirements_path: Path | None = None) -> list[PlanDiagnostic]:
    diagnostics: list[PlanDiagnostic] = []
    if path.suffix.lower() != ".xlsx":
        return [PlanDiagnostic("PLAN-001", "error", path, None, None, "test plan must use the .xlsx extension")]
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
    except (OSError, InvalidFileException, ValueError, KeyError, zipfile.BadZipFile) as exc:
        return [PlanDiagnostic("PLAN-001", "error", path, None, None, f"cannot open workbook: {exc}")]
    if "cmd" not in workbook.sheetnames:
        workbook.close()
        return [PlanDiagnostic("PLAN-002", "error", path, None, None, "missing cmd worksheet")]
    sheet = workbook["cmd"]
    actual_headers = tuple(_text(sheet.cell(1, column).value) for column in range(1, 11))
    if actual_headers != HEADERS:
        diagnostics.append(_diag(path, "PLAN-003", f"A-J headers must be: {', '.join(HEADERS)}", 1))
    for column in range(11, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            if _text(sheet.cell(row, column).value):
                diagnostics.append(_diag(path, "PLAN-003", f"undefined business data found after column J (column {column})", row))
                break

    rows: list[tuple[int, list[str]]] = []
    names_by_command: dict[str, set[str]] = {}
    indices: dict[tuple[str, str], list[int]] = {}
    searchable: dict[str, list[str]] = {}
    for row_number in range(2, sheet.max_row + 1):
        values = [_text(sheet.cell(row_number, column).value) for column in range(1, 11)]
        if not any(values[:6]):
            continue
        rows.append((row_number, values))
        row_dimension = sheet.row_dimensions[row_number]
        if row_dimension.hidden or row_dimension.collapsed:
            diagnostics.append(_diag(path, "PLAN-014", "case row must be visible and not collapsed", row_number))
        if any(not value for value in values[:6]):
            diagnostics.append(_diag(path, "PLAN-004", "case row must populate every core field A-F", row_number))
            diagnostics.append(_diag(path, "PLAN-014", "partial placeholder/history row is not allowed", row_number))
        try:
            number = int(values[0])
            if number <= 0 or str(number) != values[0]:
                raise ValueError
        except ValueError:
            diagnostics.append(_diag(path, "PLAN-005", "number must be a positive integer", row_number))
        command = values[1]
        if command and not COMMAND_NAME.fullmatch(command):
            diagnostics.append(_diag(path, "PLAN-006", "被测命令 must contain exactly one EDA command name", row_number))
        case_name = Path(values[2].replace("\\", "/")).name if values[2] else ""
        match = CASE_NAME.fullmatch(case_name)
        if not match:
            diagnostics.append(_diag(path, "PLAN-007", "case path must end in posNNN_<scenario> or negNNN_<scenario>", row_number))
        elif command:
            known = names_by_command.setdefault(command, set())
            if case_name in known:
                diagnostics.append(_diag(path, "PLAN-007", f"duplicate case name for {command}: {case_name}", row_number))
            known.add(case_name)
            indices.setdefault((command, match.group(1)), []).append(int(match.group(2)))
        description, steps, expected = values[3], values[4], values[5]
        for field_name, field_value in (("用例描述", description), ("用例步骤", steps), ("用例预期", expected)):
            if _has_inline_numbered_items(field_value):
                diagnostics.append(_diag(path, "PLAN-015", f"{field_name} must put each numbered item on its own line within the cell", row_number))
        if description and (len(description) < 12 or description in {command, f"{command}功能测试", "功能测试"}):
            diagnostics.append(_diag(path, "PLAN-008", "description must identify the scenario and covered behavior", row_number))
        if steps and (not STEP_NUMBER.search(steps) or command.lower() not in steps.lower() or not re.search(r"准备|读入|加载|创建|调用|执行|运行|检查|收集|保存|写出", steps, re.I)):
            diagnostics.append(_diag(path, "PLAN-009", "steps must be numbered and include preparation/action/result collection", row_number))
        if expected and (expected.upper() in {"PASS", "执行成功", "成功"} or not OBSERVABLE.search(expected)):
            diagnostics.append(_diag(path, "PLAN-010", "expected result must be observable and specific", row_number))
        status = values[7].upper()
        if values[6] and (len(values[6]) < 2 or values[6].upper() in {"TBD", "TODO"}):
            diagnostics.append(_diag(path, "PLAN-011", "责任人 must be a clear name", row_number, "warning"))
        if status not in VALID_STATUSES:
            diagnostics.append(_diag(path, "PLAN-011", f"unsupported status value: {values[7]}", row_number, "warning"))
        if match:
            combined = " ".join((description, expected))
            if match.group(1) == "pos" and NEGATIVE.search(combined) and not POSITIVE.search(combined):
                diagnostics.append(_diag(path, "PLAN-012", "pos case describes rejection or invalid input", row_number))
            if match.group(1) == "neg" and (not NEGATIVE.search(combined) or (POSITIVE.search(combined) and not NEGATIVE.search(expected))):
                diagnostics.append(_diag(path, "PLAN-012", "neg case must describe invalid input and expected rejection", row_number))
        if command:
            searchable.setdefault(command, []).append(" ".join((description, steps, values[8])).lower())

    for (command, polarity), found in indices.items():
        ordered = sorted(found)
        if ordered and ordered != list(range(1, len(ordered) + 1)):
            diagnostics.append(_diag(path, "PLAN-007", f"{command} {polarity} indices must start at 001 and be contiguous"))

    if rows and sheet.auto_filter.ref:
        expected_filter = f"A1:J{max(row_number for row_number, _values in rows)}"
        if sheet.auto_filter.ref != expected_filter:
            diagnostics.append(_diag(path, "PLAN-014", f"auto-filter range must match actual A-J data range: {expected_filter}", 1))

    requirements, requirement_errors = load_requirements(requirements_path)
    diagnostics.extend(requirement_errors)
    for command, points in requirements.items():
        corpus = " ".join(searchable.get(command, []))
        for point in points:
            if point.lower() not in corpus:
                diagnostics.append(_diag(path, "PLAN-013", f"uncovered test point for {command}: {point}"))
    workbook.close()
    return diagnostics


def plan_report(path: Path, requirements_path: Path | None = None) -> dict[str, object]:
    diagnostics = [item.to_dict() for item in validate_test_plan(path, requirements_path)]
    errors = [item for item in diagnostics if item["severity"] == "error"]
    return {
        "status": "FAIL" if errors else "PASS",
        "checks": [{"id": "L1", "status": "FAIL" if errors else "PASS", "input": str(path.resolve()), "diagnostics": diagnostics}],
        "diagnostics": diagnostics,
    }
