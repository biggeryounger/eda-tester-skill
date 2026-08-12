#!/usr/bin/env python3
"""Generate a non-empty EDA Excel test plan from structured scenarios."""

from __future__ import annotations

import argparse
import copy
import json
import math
import os
import re
import sys
import tempfile
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

import test_plan_validator


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT / "assets" / "templates" / "测试用例设计表.xlsx"
HEADERS = test_plan_validator.HEADERS
COLUMN_WIDTHS = {
    "A": 8,
    "B": 18,
    "C": 26,
    "D": 32,
    "E": 36,
    "F": 32,
    "G": 12,
    "H": 12,
    "I": 26,
    "J": 18,
}
MIN_DATA_ROW_HEIGHT = 30
MAX_DATA_ROW_HEIGHT = 240
LINE_HEIGHT = 18
FEATURE_TAG_PREFIX = re.compile(r"^\s*(?:\[feature\b[^\]]*\]\s*)+", re.IGNORECASE)


def sanitize_description(value: object) -> object:
    if not isinstance(value, str):
        return value
    return FEATURE_TAG_PREFIX.sub("", value).strip()


def display_width(value: object) -> int:
    return sum(2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1 for character in str(value))


def wrapped_line_count(value: object, column_width: float) -> int:
    available = max(1, int(column_width) - 2)
    logical_lines = str(value or "").splitlines() or [""]
    return sum(max(1, math.ceil(display_width(line) / available)) for line in logical_lines)


def apply_readable_layout(sheet: object, last_row: int) -> None:
    for column, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 28
    for row in sheet.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=10):
        for cell in row:
            alignment = copy.copy(cell.alignment)
            cell.alignment = Alignment(
                horizontal=alignment.horizontal,
                vertical="top" if cell.row > 1 else "center",
                text_rotation=alignment.text_rotation,
                wrap_text=True,
                shrink_to_fit=alignment.shrink_to_fit,
                indent=alignment.indent,
            )
    for row_number in range(2, last_row + 1):
        lines = max(
            wrapped_line_count(sheet.cell(row_number, column).value, COLUMN_WIDTHS[sheet.cell(row_number, column).column_letter])
            for column in range(1, 11)
        )
        sheet.row_dimensions[row_number].height = min(
            MAX_DATA_ROW_HEIGHT,
            max(MIN_DATA_ROW_HEIGHT, lines * LINE_HEIGHT + 6),
        )


def load_scenarios(path: Path) -> list[dict[str, object]]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"cannot read scenario JSON: {exc}") from exc
    scenarios = data.get("scenarios") if isinstance(data, dict) else None
    if not isinstance(scenarios, list) or not scenarios:
        raise ValueError("scenario JSON must contain at least one scenario")
    for index, scenario in enumerate(scenarios, 1):
        if not isinstance(scenario, dict):
            raise ValueError(f"scenario {index} must be an object")
        scenario["用例描述"] = sanitize_description(scenario.get("用例描述"))
        missing = [header for header in HEADERS[:6] if scenario.get(header) in (None, "")]
        if missing:
            raise ValueError(f"scenario {index} missing core fields: {', '.join(missing)}")
    return scenarios


def generate(source: Path, output: Path, template: Path = DEFAULT_TEMPLATE, requirements: Path | None = None) -> None:
    scenarios = load_scenarios(source)
    workbook = load_workbook(template, data_only=False)
    if "cmd" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("template is missing cmd worksheet")
    sheet = workbook["cmd"]
    if tuple(sheet.cell(1, column).value for column in range(1, 11)) != HEADERS:
        workbook.close()
        raise ValueError("template A-J headers do not match the required structure")
    style_cells = [copy.copy(sheet.cell(2, column)._style) for column in range(1, 11)]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for row_number, scenario in enumerate(scenarios, 2):
        for column, header in enumerate(HEADERS, 1):
            cell = sheet.cell(row_number, column, scenario.get(header, ""))
            cell._style = copy.copy(style_cells[column - 1])
        sheet.row_dimensions[row_number].hidden = False
        sheet.row_dimensions[row_number].collapsed = False
        sheet.row_dimensions[row_number].outlineLevel = 0
    sheet.auto_filter.ref = f"A1:J{len(scenarios) + 1}"
    apply_readable_layout(sheet, len(scenarios) + 1)
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary_name: str | None = None
    try:
        with tempfile.NamedTemporaryFile(prefix="eda-plan-", suffix=".xlsx", dir=output.parent, delete=False) as temporary:
            temporary_name = temporary.name
        workbook.save(temporary_name)
        workbook.close()
        generated = Path(temporary_name)
        reopened = load_workbook(generated, read_only=True, data_only=False)
        generated_sheet = reopened["cmd"]
        populated = sum(1 for row in generated_sheet.iter_rows(min_row=2, max_col=6, values_only=True) if all(value not in (None, "") for value in row))
        reopened.close()
        if populated != len(scenarios):
            raise ValueError(f"generated workbook contains {populated} complete rows; expected {len(scenarios)}")
        diagnostics = test_plan_validator.validate_test_plan(generated, requirements)
        errors = [item for item in diagnostics if item.severity == "error"]
        if errors:
            summary = "; ".join(f"{item.rule_id}: {item.message}" for item in errors[:5])
            raise ValueError(f"generated workbook failed Layer 1: {summary}")
        os.replace(generated, output)
        temporary_name = None
    finally:
        workbook.close()
        if temporary_name:
            Path(temporary_name).unlink(missing_ok=True)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("scenarios", type=Path, help="JSON file containing a non-empty scenarios array")
    parser.add_argument("output", type=Path)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--requirements", type=Path)
    args = parser.parse_args()
    try:
        generate(args.scenarios.resolve(), args.output.resolve(), args.template.resolve(), args.requirements.resolve() if args.requirements else None)
    except (OSError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(f"Generated {args.output} with validated, non-empty test cases.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
